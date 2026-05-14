import customtkinter as ctk
from openpyxl import load_workbook
import os
from tkinter import messagebox
import datetime

class DashboardWindow(ctk.CTkToplevel):
    def __init__(self, parent, arquivo_path):
        super().__init__(parent)
        self.title("Dashboard de Controle e Relatórios")
        self.geometry("750x750")
        self.arquivo_path = arquivo_path
        
        # Caminhos fixos do projeto
        self.diretorio_anexos = r"C:\Users\user\Desktop\Projeto pi\anexo"
        self.diretorio_relatorios = r"C:\Users\user\Desktop\Projeto pi\relatorio_mes"
        
        if not os.path.exists(self.diretorio_relatorios):
            os.makedirs(self.diretorio_relatorios)
            
        self.attributes("-topmost", True)
        self.criar_layout()
        self.carregar_dados_iniciais()

    def criar_layout(self):
        ctk.CTkLabel(self, text="📊 Dashboard de Indicadores", font=("Arial", 22, "bold")).pack(pady=15)
        
        # --- SEÇÃO DE FILTRO ---
        frame_busca = ctk.CTkFrame(self)
        frame_busca.pack(pady=10, padx=20, fill="x")
        
        ctk.CTkLabel(frame_busca, text="Mês do Relatório:", font=("Arial", 12)).pack(side="left", padx=10)
        self.meses = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
        self.combo_mes = ctk.CTkComboBox(frame_busca, values=self.meses, width=100)
        self.combo_mes.set("05") 
        self.combo_mes.pack(side="left", padx=5)
        
        btn_gerar = ctk.CTkButton(frame_busca, text="🔍 Gerar Relatório .txt", fg_color="#d35400", 
                                 hover_color="#e67e22", command=self.processar_filtro_data)
        btn_gerar.pack(side="left", padx=10, pady=10)

        # --- CARDS DE INDICADORES ---
        self.frame_cards = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_cards.pack(pady=10, padx=20, fill="x")
        self.card_total = self.criar_card(self.frame_cards, "Total de Registros", "0")
        self.card_participantes = self.criar_card(self.frame_cards, "Participantes (Geral)", "0")

        # --- LISTA DE ANEXOS ---
        ctk.CTkLabel(self, text="📂 Arquivos na Pasta Anexo", font=("Arial", 16, "bold")).pack(pady=(20, 5))
        self.frame_anexos = ctk.CTkScrollableFrame(self, width=650, height=250)
        self.frame_anexos.pack(pady=10, padx=20, fill="both", expand=True)

    def criar_card(self, master, titulo, valor):
        f = ctk.CTkFrame(master)
        f.pack(side="left", padx=10, expand=True, fill="both")
        ctk.CTkLabel(f, text=titulo).pack(pady=(5,0))
        lbl_val = ctk.CTkLabel(f, text=valor, font=("Arial", 22, "bold"), text_color="#1f538d")
        lbl_val.pack(pady=10)
        return lbl_val

    def processar_filtro_data(self):
        if not self.arquivo_path:
            messagebox.showerror("Erro", "Arquivo Excel não encontrado!")
            return

        mes_alvo = int(self.combo_mes.get())
        
        try:
            wb = load_workbook(self.arquivo_path, data_only=True)
            ws = wb.active
            
            # MAPEAMENTO FIXO BASEADO NA IMAGEM
            IDX_SEC = 2      # Coluna C
            IDX_DATA = 13    # Coluna N
            IDX_PART = 14    # Coluna O
            IDX_CONTE = 18   # Coluna S
            IDX_VISITA = 19  # Coluna T
            IDX_TEMPO = 20   # Coluna U

            dados_do_mes = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                valor_data = row[IDX_DATA]
                if not valor_data: continue

                eh_mes_correto = False
                if isinstance(valor_data, (datetime.datetime, datetime.date)):
                    if valor_data.month == mes_alvo: eh_mes_correto = True
                else:
                    d_str = str(valor_data)
                    if f"/{mes_alvo:02d}/" in d_str or f"/{mes_alvo}/" in d_str: eh_mes_correto = True

                if eh_mes_correto:
                    info = {
                        "secretaria": str(row[IDX_SEC]) if row[IDX_SEC] else "Não Informado",
                        "participantes": int(row[IDX_PART]) if str(row[IDX_PART]).isdigit() else 0,
                        "avaliacoes": [
                            str(row[IDX_CONTE]).strip() if row[IDX_CONTE] else "",
                            str(row[IDX_VISITA]).strip() if row[IDX_VISITA] else "",
                            str(row[IDX_TEMPO]).strip() if row[IDX_TEMPO] else ""
                        ]
                    }
                    dados_do_mes.append(info)

            if not dados_do_mes:
                messagebox.showinfo("Aviso", f"Sem dados para o mês {mes_alvo:02d}")
                return

            self.escrever_txt(mes_alvo, dados_do_mes)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro no processamento: {e}")

    def escrever_txt(self, mes, dados):
        caminho_txt = os.path.join(self.diretorio_relatorios, f"Relatorio_Mensal_{mes:02d}.txt")
        total_p = sum(d["participantes"] for d in dados)
        secretarias = sorted(set(d["secretaria"] for d in dados))
        
        contagem_votos = {"Ótimo": 0, "Bom": 0, "Regular": 0, "Ruim": 0}
        for d in dados:
            for v in d["avaliacoes"]:
                if v in contagem_votos: contagem_votos[v] += 1

        try:
            # utf-8-sig resolve o problema de acentuação no Bloco de Notas
            with open(caminho_txt, "w", encoding="utf-8-sig") as f:
                f.write(f"RELATÓRIO DE ATIVIDADES - MÊS {mes:02d}/2026\n")
                f.write("="*45 + "\n")
                f.write(f"Total de Ações Registradas: {len(dados)}\n")
                f.write(f"Total de Participantes Atendidos: {total_p}\n")
                f.write("-" * 45 + "\n\n")
                f.write("Secretarias/Autarquias Atendidas:\n")
                for s in secretarias: f.write(f" - {s}\n")
                f.write("\nResumo das Avaliações (Geral):\n")
                f.write("(Conteúdo, Visita e Tempo Dedicado)\n")
                total_votos = sum(contagem_votos.values())
                for k, v in contagem_votos.items():
                    p = (v/total_votos*100) if total_votos > 0 else 0
                    f.write(f"  {k.ljust(10)}: {str(v).rjust(2)} votos ({p:.1f}%)\n")
                f.write("\n" + "="*45 + "\n")

            os.startfile(caminho_txt)
            messagebox.showinfo("Sucesso", "Relatório gerado com acentuação corrigida!")
        except Exception as e:
            messagebox.showerror("Erro ao salvar", str(e))

    def carregar_dados_iniciais(self):
        if self.arquivo_path:
            try:
                wb = load_workbook(self.arquivo_path, data_only=True)
                ws = wb.active
                self.card_total.configure(text=str(max(0, ws.max_row - 1)))
                soma_geral = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > 14 and str(row[14]).isdigit():
                        soma_geral += int(row[14])
                self.card_participantes.configure(text=str(soma_geral))
            except: pass

        for w in self.frame_anexos.winfo_children(): w.destroy()
        if os.path.exists(self.diretorio_anexos):
            for arq in os.listdir(self.diretorio_anexos):
                caminho_f = os.path.join(self.diretorio_anexos, arq)
                if os.path.isfile(caminho_f):
                    item = ctk.CTkFrame(self.frame_anexos)
                    item.pack(fill="x", pady=2, padx=5)
                    ctk.CTkLabel(item, text=arq, anchor="w").pack(side="left", padx=10, expand=True)
                    ctk.CTkButton(item, text="Abrir", width=60, height=24, fg_color="#27ae60",
                                 command=lambda p=caminho_f: os.startfile(p)).pack(side="right", padx=5)