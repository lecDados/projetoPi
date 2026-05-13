import customtkinter as ctk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import os

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

colunas = [
    "Turmas", "Responsável", "Secretaria/autarquia", "E-mail",
    "Endereço", "Participantes", "Autorização", "Lista de presença",
    "Publicações - Instagram", "Publicações - Facebook", "Publicações - Outras"
]

arquivo_path = None


app = ctk.CTk()
app.title("Cadastro Excel")
app.geometry("600x700")

entries = {}


def mostrar_sucesso():
    popup = ctk.CTkToplevel(app)
    popup.title("Sucesso")
    popup.geometry("300x150")

    ctk.CTkLabel(
        popup,
        text="✅ Enviado com sucesso!",
        font=("Arial", 16)
    ).pack(pady=30)

    ctk.CTkButton(
        popup,
        text="OK",
        command=popup.destroy
    ).pack(pady=10)


def selecionar_arquivo():
    global arquivo_path
    arquivo_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if arquivo_path:
        label_arquivo.configure(text=f"Arquivo: {os.path.basename(arquivo_path)}")


def salvar_dados():
    if not arquivo_path:
        print("Selecione um arquivo primeiro")
        return

    try:
        wb = load_workbook(arquivo_path)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active

    
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        for i, col in enumerate(colunas, start=1):
            ws.cell(row=1, column=i, value=col)

    
    linha = ws.max_row + 1

    
    for i, col in enumerate(colunas, start=1):
        valor = entries[col].get()
        ws.cell(row=linha, column=i, value=valor)

    wb.save(arquivo_path)

    
    for entry in entries.values():
        entry.delete(0, 'end')

    
    mostrar_sucesso()


ctk.CTkLabel(app, text="Sistema de Cadastro Excel", font=("Arial", 20)).pack(pady=10)

btn_arquivo = ctk.CTkButton(app, text="Selecionar Arquivo Excel", command=selecionar_arquivo)
btn_arquivo.pack(pady=10)

label_arquivo = ctk.CTkLabel(app, text="Nenhum arquivo selecionado")
label_arquivo.pack(pady=5)

frame = ctk.CTkScrollableFrame(app, width=550, height=400)
frame.pack(pady=10)


for col in colunas:
    label = ctk.CTkLabel(frame, text=col)
    label.pack(anchor="w", padx=10)

    entry = ctk.CTkEntry(frame, width=500)
    entry.pack(padx=10, pady=5)

    entries[col] = entry


btn_salvar = ctk.CTkButton(app, text="Salvar no Excel", command=salvar_dados)
btn_salvar.pack(pady=20)

app.mainloop()