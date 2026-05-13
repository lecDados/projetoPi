from openpyxl import load_workbook, Workbook
from tkinter import messagebox


def salvar_no_excel(arquivo_path, dados, colunas):

    try:

        wb = load_workbook(arquivo_path)
        ws = wb.active

    except:

        wb = Workbook()
        ws = wb.active

    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:

        for i, col in enumerate(colunas, start=1):

            ws.cell(
                row=1,
                column=i,
                value=col
            )

    linha = ws.max_row + 1

    for i, col in enumerate(colunas, start=1):

        ws.cell(
            row=linha,
            column=i,
            value=dados.get(col)
        )

    try:

        wb.save(arquivo_path)

    except PermissionError:

        messagebox.showerror(
            "Erro",
            "Feche o arquivo Excel antes de salvar."
        )